#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从 test_summary.tsv 同步最新测试状态到 Excel（迁移自 dongwei/scripts/update_xlsx.py）。

优先级规则（固化）：history PASS > 真实 PASS > 其他；已有 PASS 后失败不覆盖；DEVICE_OFFLINE 不覆盖。
SKIP 辅助包在主包通过后也判 PR 满足（Test.json kits 映射）。
"""
from __future__ import annotations

import json
import os
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from common.paths import REPO, TSV, XLSX  # noqa: E402

from openpyxl import load_workbook  # noqa: E402


def _better_status(cur: tuple, status: str, passed: str, total: str, err: str,
                    report: str, line: str) -> tuple | None:
    """PASS 优先状态合并；返回更优记录或 None（保留旧值）。"""
    if cur[0] == "PASS" and line.startswith("history"):
        return None
    if status == "PASS":
        return (status, passed, total, err, report)
    if cur[0] == "PASS" or status == "DEVICE_OFFLINE":
        return None
    return (status, passed, total, err, report)


def latest_from_tsv() -> dict:
    """每工程最新状态（history PASS 最高优先级）。"""
    latest: dict[str, tuple] = {}
    for line in TSV.read_text(errors="replace").splitlines():
        parts = line.rstrip("\n").split("\t")
        if len(parts) < 7:
            continue
        ts, rel, status, passed, total, err, report = parts[:7]
        if rel not in latest:
            latest[rel] = (status, passed, total, err, report)
            continue
        new = _better_status(latest[rel], status, passed, total, err, report, line)
        if new:
            latest[rel] = new
    return latest


def _load_test_json(root: str) -> dict | None:
    """读取目录下 Test.json（失败返回 None）。"""
    try:
        return json.load(open(os.path.join(root, "Test.json")))
    except Exception:
        return None


def _kits_to_map(d: dict, root: str, hap_to_main: dict) -> None:
    """AppInstallKit 的 test-file-name 注册到 hap 名 → 工程路径映射。"""
    for kit in d.get("kits", []):
        if kit.get("type") != "AppInstallKit":
            continue
        for hap in kit.get("test-file-name", []):
            hap_to_main.setdefault(os.path.splitext(hap)[0], []).append(
                os.path.relpath(root, str(REPO)))


def build_hap_to_main_map() -> dict[str, list[str]]:
    """辅助包 hap 名 → 主工程相对路径（Test.json kits AppInstallKit）。"""
    hap_to_main: dict[str, list[str]] = {}
    for root, dirs, files in os.walk(str(REPO / "ability")):
        if "oh_modules" in root or "/build/" in root:
            continue
        if "Test.json" in files:
            d = _load_test_json(root)
            if d:
                _kits_to_map(d, root, hap_to_main)
    return hap_to_main


def _sync_status(row, latest) -> bool:
    """状态/通过数/报告/备注同步。返回是否更新。"""
    p = row[2].value
    if p not in latest:
        return False
    status, passed, total, err, report = latest[p]
    tst = {"PASS": "通过", "FAIL": "失败", "BUILD_FAIL": "编译失败",
           "INSTALL_FAIL": "安装失败", "DEVICE_OFFLINE": "设备离线",
           "NO_RESULT": "无结果", "SIGN_FAIL": "签名失败"}.get(status, status)
    row[6].value = tst
    row[7].value = f"{passed}/{total}" if passed != "-" else ""
    row[9].value = report if report != "-" else ""
    if status == "PASS":
        if row[10].value and str(row[10].value).startswith("[auto]"):
            row[10].value = None
    elif err and err != "-":
        _update_note(row, err)
    return True


def _update_note(row, err: str) -> None:
    """失败备注写入（保留人工判定/主hap未通过等既有备注）。"""
    note = str(err)[:200]
    cur_note = str(row[10].value or "")
    if any(k in cur_note for k in ("已判定限制", "主hap暂未测试通过", "主包(")):
        return
    if cur_note and not cur_note.startswith("[auto]"):
        row[10].value = f"{cur_note}；[auto]{note}"
    else:
        row[10].value = f"[auto]{note}"




def _resolve_skip_pr(p: str, hap_to_main: dict, main_status: dict, row) -> str:
    """SKIP 辅助 hap：主 hap 已通过 → 满足 PR；未通过则备注说明。"""
    name = os.path.basename(p).lower()
    main_names: list[str] = []
    for hap, mains in hap_to_main.items():
        if name in hap.lower() or hap.lower() in name:
            main_names = mains
            if any(main_status.get(m, ("", ""))[0] == "通过" for m in mains):
                return "是"
    if main_names:
        note = "主hap暂未测试通过：" + "、".join(m.rsplit("/", 1)[-1] for m in main_names)
        cur = row[10].value or ""
        row[10].value = f"{cur}；[auto]{note}" if not str(cur).startswith("[auto]") else f"[auto]{note}"
    return "否"


def sync_xlsx() -> int:
    latest = latest_from_tsv()
    wb = load_workbook(XLSX)
    ws = wb["需求1进度总表"]
    updated = 0
    main_status = {}
    for row in ws.iter_rows(min_row=2):
        p = row[2].value or ""
        if p.startswith("ability/"):
            main_status[p] = (row[6].value, row[8].value)
    hap_to_main = build_hap_to_main_map()
    for row in ws.iter_rows(min_row=2):
        p = row[2].value
        if not _sync_status(row, latest):
            continue
        # [manual] 手工判定保护：备注含"已通过"判是，否则保持
        if str(row[10].value or "").startswith("[manual]"):
            if "已通过" in str(row[10].value):
                row[8].value = "是"
            continue
        comp = row[4].value
        tst = row[6].value
        pr = "是" if (comp == "编译通过" and tst == "通过") else "否"
        if pr == "否" and tst == "SKIP" and p.startswith("ability/"):
            pr = _resolve_skip_pr(p, hap_to_main, main_status, row)
        row[8].value = pr
        updated += 1
    wb.save(XLSX)
    return updated





def main() -> None:
    print(f"Excel 已更新 {sync_xlsx()} 行")


if __name__ == "__main__":
    main()
