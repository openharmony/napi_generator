#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""批量逐个闭环调度（7.6 模式）：表格顺序 → 逐个 HAP 测试 → 失败即停分析修复。

规则（固化 SKILL 7.1/7.6）：
- 必须逐个 HAP 闭环：一个调试到通过（或明确判定限制）后才处理下一个
- 表格（test_summary.tsv）中 history PASS 跳过不反复测试
- 每 3 分钟刷新进度快照（progress_report.sh）+ 同步 Excel
- 失败即停：任一个失败 → 调 triage 处置 → 不自动进入下一个
"""
from __future__ import annotations

import argparse
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from common.paths import PROGRESS_DIR, TSV  # noqa: E402
from report.progress_report import refresh_snapshot  # noqa: E402

SKIP_STATUSES = {"PASS", "SKIP"}
VERDICT_LIST = Path("/tmp/xts_skip_verdicts.txt")   # 已判定限制的跳过清单（每行 rel\t原因）


def load_verdict_list() -> dict[str, str]:
    """加载已判定限制清单：rel → 原因（人工/自动判定后写入，批量直接跳过）。"""
    d: dict[str, str] = {}
    try:
        for line in VERDICT_LIST.read_text(errors="replace").splitlines():
            parts = line.split("\t")
            if len(parts) >= 2 and parts[0].strip():
                d[parts[0].strip()] = parts[1].strip()
    except OSError:
        pass
    return d


def table_seq_map() -> dict[str, int]:
    """需求1进度表.xlsx 的 工程路径→序号 映射（输出以表格序号为准，第 1 点）。"""
    m: dict[str, int] = {}
    try:
        from openpyxl import load_workbook
        wb = load_workbook('/root/aiSkill/develop/dongwei/进度/需求1进度表.xlsx', read_only=True)
        ws = wb['需求1进度总表']
        for row in ws.iter_rows(min_row=2, max_col=3):
            seq, _, rel = row[0].value, row[1].value, row[2].value
            if seq and rel:
                m[str(rel)] = int(seq)
    except Exception:
        pass
    return m


def history_passed(rel: str) -> bool:
    """TSV 中该工程是否有任意 PASS 记录（history PASS 最高优先级，不反复测试）。"""
    try:
        for line in TSV.read_text(errors="replace").splitlines():
            parts = line.split("\t")
            if len(parts) >= 3 and parts[1] == rel and parts[2] == "PASS":
                return True
    except OSError:
        pass
    return False


def test_one(rel: str, profile: str) -> int:
    """调 test/test_one.py，返回退出码。"""
    script = Path(__file__).resolve().parent / "test_one.py"
    args = [sys.executable, str(script), rel]
    if profile != "release":
        args += ["--profile", profile]
    r = subprocess.run(args, capture_output=False)
    return r.returncode


def last_err(rel: str) -> str:
    """该工程 TSV 最新一行的 err 列。"""
    err = ""
    try:
        for line in TSV.read_text(errors="replace").splitlines():
            parts = line.split("\t")
            if len(parts) >= 6 and parts[1] == rel:
                err = parts[5]
    except OSError:
        pass
    return err


def record_verdict(rel: str, err: str, hint: str, rule: str) -> None:
    """更新备注：追加 失败分析记录.md + 判定日志（跳过继续的落盘点）。"""
    from datetime import datetime
    entry = (f"\n### [{rel}] {datetime.now():%Y-%m-%d %H:%M:%S}（批量闭环判定，跳过继续）\n"
             f"- 现象: {err[:200]}\n"
             f"- 方案: {rule}\n"
             f"- 判定: 设备/环境限制（{hint[:120]}）\n"
             f"- 处理: 标记限制，不阻塞后续\n")
    try:
        with open(PROGRESS_DIR / "失败分析记录.md", "a") as f:
            f.write(entry)
    except OSError:
        pass
    try:
        with open("/tmp/xts_verdicts.tsv", "a") as f:
            f.write(f"{datetime.now():%Y%m%d_%H%M%S}\t{rel}\t{rule}\t{err[:120]}\n")
    except OSError:
        pass


def run_batch(rel_list: Path | None, profile: str, skip_passed: bool = True,
              verdict_skip: bool = False) -> None:
    log = PROGRESS_DIR / "batch_test.log"
    items = [ln.strip() for ln in rel_list.read_text().splitlines() if ln.strip()] if rel_list else []
    # 批量开始前全局清理设备（残留应用多会造成包名冲突/脏 log，2026-08-18 强化）
    from common import hdc_utils as _hdc
    if _hdc.ensure_device():
        _hdc.cleanup_device()
        print("[batch] 全局清理完成（卸载全部非系统应用）")
    total = len(items)
    seq_map = table_seq_map()
    def tag(rel: str) -> str:
        return f"表格#{seq_map.get(rel, '?')}" if seq_map else f"[{i}/{total}]"
    done = passed = skipped = 0
    with open(log, "a") as f:
        f.write(f"\n===== BATCH {datetime.now():%Y-%m-%d %H:%M:%S} ({total} 工程) =====\n")
    for i, rel in enumerate(items, 1):
        if skip_passed and history_passed(rel):
            with open(log, "a") as f:
                f.write(f"[{i}/{total}] SKIP(history PASS) {rel}\n")
            print(f"{tag(rel)} SKIP(history PASS) {rel}")
            done += 1
            skipped += 1
            continue
        # 已判定限制清单：直接跳过（更新备注一次）
        verdict_reasons = load_verdict_list()
        if rel in verdict_reasons:
            record_verdict(rel, f"已判定限制: {verdict_reasons[rel]}",
                           verdict_reasons[rel], "skip-list")
            with open(log, "a") as f:
                f.write(f"[{i}/{total}] VERDICT_SKIP {rel} ({verdict_reasons[rel]})\n")
            print(f"⏭️ {tag(rel)} {rel} 已判定限制（{verdict_reasons[rel]}），跳过")
            done += 1
            skipped += 1
            continue
        print(f"\n{tag(rel)} ▶ {rel}")
        with open(log, "a") as f:
            f.write(f"[{i}/{total}] TEST {rel}\n")
        rc = test_one(rel, profile)
        done += 1
        if rc == 0:
            passed += 1
            with open(log, "a") as f:
                f.write(f"DONE {rel}\n")
        elif rc in (2, 3):
            # 设备离线(2)/辅助工程无测试(3)：记录后继续（非真实失败）
            skipped += 1
            with open(log, "a") as f:
                f.write(f"SKIP_ITEM {rel} rc={rc}\n")
            print(f"{tag(rel)} {rel} 跳过（rc={rc}：设备离线/辅助工程）")
        else:
            # 失败：--verdict-skip 模式下先判定（方案库 verdict=limit → 更新备注跳过继续）
            if verdict_skip:
                err = last_err(rel)
                from triage import dispatch
                v = dispatch(rel, err)
                if v.get("verdict") == "limit":
                    record_verdict(rel, err, v.get("hint", ""), v.get("rule", ""))
                    skipped += 1
                    with open(log, "a") as f:
                        f.write(f"VERDICT_SKIP {rel} ({v.get('rule')})\n")
                    print(f"⏭️ [{i}/{total}] {rel} 设备/环境限制（{v.get('rule')}），"
                          f"已更新备注，跳过继续")
                    if i % 3 == 0:
                        refresh_snapshot()
                    time.sleep(1)
                    continue
            with open(log, "a") as f:
                f.write(f"FAIL_STOP {rel} rc={rc}\n")
            print(f"⏸️ {tag(rel)} {rel} 失败（rc={rc}），按 7.6 逐个闭环：triage → 修复 → 重测")
            print(f"   → python3 scripts/test/triage.py {rel}")
            break
        if i % 3 == 0:
            refresh_snapshot()
        time.sleep(1)
    refresh_snapshot()
    print(f"\n===== BATCH DONE: passed={passed} skipped={skipped} stopped_at={done}/{total} =====")


def pending_list(subdir: str = "") -> tuple[list[str], dict]:
    """从 TSV 生成未闭环清单（表格顺序）+ 状态统计。

    判定规则（与 update_xlsx 一致的 PASS 优先语义）：
    - PASS 永不被后续记录覆盖；SKIP/DEVICE_OFFLINE 可被后续真实状态覆盖
    - 已闭环 = 最新状态 ∈ {PASS, SKIP}；未闭环 = 其余全部
    """
    latest: dict[str, str] = {}
    order: list[str] = []
    for line in TSV.read_text(errors="replace").splitlines():
        parts = line.split("\t")
        if len(parts) < 3:
            continue
        rel, st = parts[1], parts[2]
        if rel not in latest:
            order.append(rel)
            latest[rel] = st
        elif st == "PASS" and latest[rel] != "PASS":
            latest[rel] = st
        elif st not in ("DEVICE_OFFLINE",) and latest[rel] in ("PASS", "SKIP", "DEVICE_OFFLINE"):
            latest[rel] = st
    items = [r for r in order if latest[r] not in SKIP_STATUSES]
    if subdir:
        items = [r for r in items if r.startswith(subdir)]
    from collections import Counter
    stats = {"total": len(latest), "closed": sum(1 for s in latest.values() if s in SKIP_STATUSES),
             "open": len(items)}
    stats["by_status"] = dict(Counter(latest[r] for r in items))
    return items, stats


def main() -> None:
    ap = argparse.ArgumentParser(description="批量逐个闭环测试")
    ap.add_argument("--list", default="", help="工程清单文件（每行一个相对路径）；缺省从 TSV 未 PASS 项生成")
    ap.add_argument("--subdir", default="ability/ability_runtime", help="--list 缺省时的扫描范围")
    ap.add_argument("--profile", choices=["release", "debug", "system"], default="release")
    ap.add_argument("--no-skip-passed", action="store_true", help="不跳过 history PASS")
    ap.add_argument("--stats", action="store_true", help="只输出未闭环判定统计，不运行")
    ap.add_argument("--verdict-skip", action="store_true",
                    help="失败项先方案库判定：设备/环境限制类自动更新备注并跳过继续")
    args = ap.parse_args()

    if args.stats:
        items, stats = pending_list(args.subdir)
        print(f"全仓 {stats['total']} 工程 | 已闭环 {stats['closed']} | 未闭环 {stats['open']}")
        print("未闭环状态构成:", stats["by_status"])
        return

    if args.list:
        rel_list = Path(args.list)
    else:
        # 从 TSV 取未闭环项（按表格行序）
        items, _ = pending_list(args.subdir)
        if not items:
            print("无未闭环工程（全部 PASS/SKIP）")
            return
        rel_list = Path("/tmp/xts_pending.txt")
        rel_list.write_text("\n".join(items) + "\n")
        print(f"生成未闭环清单 {len(items)} 项: {rel_list}")
    run_batch(rel_list, args.profile, not args.no_skip_passed, args.verdict_skip)


if __name__ == "__main__":
    main()
